# shared/strategy_auto_tuning/engine.py
"""
Strategy Auto-Tuning — Engine (V1)

End-to-end Random Search:
- Build space from TUNING
- Sample N trials
- Mutate only TUNING -> trial config xlsx
- Evaluate via run_strategia.py (StrategyEvaluator)
- Compute objective score
- Record trials.csv
- Save best.xlsx
"""

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
import csv
import json
import re
import shutil

import pandas as pd
from openpyxl import load_workbook




from .evaluator import StrategyEvaluator
from .mutator import write_trial_config
from .space import build_space_from_tuning
from .samplers import RandomSearchSampler
from .objectives import ObjectiveSpec, ObjectiveResult, compute_objective
from .regime_forward_engine import (
    EvalResult,
    ForwardSelectionSpec,
    RegimeBlock,
    run_regime_forward_selection,
)

from .io_config import (
    read_tuning_sheet,
    write_tuning_sheet,
    build_regime_wise_config_v3,
)


@dataclass(frozen=True)
class RunSpec:
    input_csv: str
    config_strategy: str
    timeframe: str
    trials: int
    seed: int
    outdir: str
    n_min_trades: int
    active_group: Optional[str] = None


@dataclass(frozen=True)
class RegimeRunSpec:
    input_csv: str
    config_strategy: str
    timeframe: str
    trials_per_regime: int
    seed: int
    outdir: str
    n_min_trades: int
    train_ratio: float
    detected_entry_groups: List[str]
    target_groups: List[str]

@dataclass(frozen=True)
class TrialRow:
    trial_id: int
    score: float
    alpha_vs_buyhold: Optional[float]
    penalty: float
    ok: bool
    reason: Optional[str]
    profit: Optional[float]
    profit_per_trade: Optional[float]
    trade_count: Optional[int]
    max_drawdown: Optional[float]
    buy_hold_profit: Optional[float]
    trial_config_xlsx: str
    eval_dir: str
    evaluation_semantics: Optional[str] = None
    gate_check_ok: Optional[bool] = None
    bad_entry_count: Optional[int] = None

def _read_trade_profit_list_from_signal_dir(signal_dir: object) -> list[float]:
    if signal_dir is None:
        return []

    try:
        p = Path(str(signal_dir))
    except Exception:
        return []

    if not p.exists():
        return []

    if p.is_file():
        candidates = [p]
    else:
        candidates = [
            x for x in p.rglob("SIGNAL_*.csv")
            if x.is_file() and not x.name.startswith("TRADE_FREQ_")
        ]

    if not candidates:
        return []

    candidates = sorted(candidates, key=lambda x: x.stat().st_mtime, reverse=True)
    signal_csv = candidates[0]

    try:
        df = pd.read_csv(signal_csv, sep=";", engine="python")
    except Exception:
        return []

    if "Profit/Trade" not in df.columns:
        return []

    vals = []
    for v in df["Profit/Trade"].tolist():
        num = _to_float_maybe(v)
        if num is not None:
            vals.append(num)

    return vals

def _fmt2(value: object) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass

    if isinstance(value, str):
        s = value.strip().lower()
        if s in {"", "nan", "none", "null"}:
            return ""
        num = _to_float_maybe(value)
        if num is None:
            return value
        return f"{num:.2f}".replace(".", ",")

    try:
        v = float(value)
    except Exception:
        return str(value)
    return f"{v:.2f}".replace(".", ",")


def _to_float_maybe(value: object) -> float | None:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass

    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return None

    try:
        return float(value)
    except Exception:
        return None

def _safe_profit_per_trade(profit: object, trade_count: object) -> float | None:
    p = _to_float_maybe(profit)
    n = _to_float_maybe(trade_count)

    if p is None or n is None:
        return None

    try:
        n_int = int(n)
    except Exception:
        return None

    if n_int <= 0:
        return None

    return p / n_int

def _read_trade_profit_list_from_best_xlsx(best_xlsx: object) -> list[float]:
    """
    Resolve trade profits for a per-regime best block.

    Cases handled:
    - best_xlsx = .../regime_G_XXX/best.xlsx
    - best_xlsx = .../regime_G_XXX/trials/trial_000N.xlsx

    In the tuned-trial case, SIGNAL artifacts usually live under sibling folders
    such as eval/ or eval_baseline/, not under trials/.
    """
    if best_xlsx is None:
        return []

    try:
        best_path = Path(str(best_xlsx))
    except Exception:
        return []

    if not best_path.exists():
        return []

    search_roots: list[Path] = []

    def _add_root(p: Path | None) -> None:
        if p is None:
            return
        try:
            pp = Path(p)
        except Exception:
            return
        if not pp.exists():
            return
        if pp not in search_roots:
            search_roots.append(pp)

    parent = best_path.parent
    _add_root(parent)

    # If best_xlsx is inside .../trials/trial_XXXX.xlsx, the real SIGNALs are
    # typically under sibling eval/ or eval_baseline/ folders at regime-root level.
    if parent.name.lower() == "trials":
        regime_root = parent.parent
        _add_root(regime_root)
        _add_root(regime_root / "eval")
        _add_root(regime_root / "eval_baseline")

    # Generic fallback: walk ancestors and prefer any regime_* folder plus its
    # common evaluation subfolders.
    for anc in best_path.parents:
        name = anc.name.lower()
        if name.startswith("regime_"):
            _add_root(anc)
            _add_root(anc / "eval")
            _add_root(anc / "eval_baseline")

    for root in search_roots:
        vals = _read_trade_profit_list_from_signal_dir(root)
        if vals:
            return vals

    return []


def _fmt_trade_list(values: list[float]) -> str:
    if not values:
        return ""
    return ", ".join(_fmt2(v) for v in values)

def _decode_regime_code(value: object) -> str:
    v = _to_float_maybe(value)
    if v is None:
        return ""

    code = int(v)
    mapping = {
        0: "LATERAL",
        1: "RANGE",
        2: "VOLATILE",
        3: "TREND_UP",
        4: "TREND_DOWN",
    }
    return mapping.get(code, str(code))


def _group_to_regime_name(group_name: str) -> str:
    mapping = {
        "G_RANGE": "RANGE",
        "G_LATERAL": "LATERAL",
        "G_VOLATILE": "VOLATILE",
        "G_TREND_UP": "TREND_UP",
        "G_TREND_DOWN": "TREND_DOWN",
        "G_UNKNOWN": "UNKNOWN",
    }
    g = str(group_name).strip().upper()
    return mapping.get(g, g.replace("G_", "", 1))


def _pick_first_existing_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lut = {str(c).strip().lower(): c for c in df.columns}
    for name in candidates:
        key = str(name).strip().lower()
        if key in lut:
            return lut[key]
    return None


def _read_signal_csv_auto(signal_csv: Path) -> pd.DataFrame:
    signal_csv = Path(signal_csv)
    try:
        return pd.read_csv(signal_csv, sep=";", engine="python")
    except Exception:
        return pd.read_csv(signal_csv)


def _verify_entries_only_on_target_regime(
    *,
    signal_csv: Path,
    target_regime_name: str,
) -> dict[str, Any]:
    """
    v3 gate-check:
    every HOLD OUT -> IN transition must occur on target regime.
    Exits are intentionally unconstrained.
    """
    signal_csv = Path(signal_csv)
    if not signal_csv.exists():
        return {
            "ok": False,
            "reason": "signal_csv_not_found",
            "entry_count": 0,
            "bad_entry_count": 0,
            "violations": [],
        }

    df = _read_signal_csv_auto(signal_csv)
    df.columns = [str(c).strip() for c in df.columns]

    hold_col = _pick_first_existing_column(df, ["HOLD", "hold", "Hold"])

    regime_col = _pick_first_existing_column(
        df,
        ["REGIME_L1", "REGIME_L1_RAW", "REGIME_L1_CODE"],
    )

    if hold_col is None:
        return {
            "ok": False,
            "reason": "missing_hold_column",
            "entry_count": 0,
            "bad_entry_count": 0,
            "violations": [],
        }

    if regime_col is None:
        return {
            "ok": False,
            "reason": "missing_regime_column",
            "entry_count": 0,
            "bad_entry_count": 0,
            "violations": [],
        }

    hold_series = df[hold_col].astype(str).str.strip().str.upper()
    prev_hold = hold_series.shift(1).fillna("OUT")
    curr_hold = hold_series.fillna("OUT")

    target_name = str(target_regime_name).strip().upper()

    regime_raw = df[regime_col]

    if str(regime_col).strip().upper() == "REGIME_L1_CODE":
        regime_series = regime_raw.map(_decode_regime_code).astype(str).str.strip().str.upper()
    else:
        regime_series = regime_raw.astype(str).str.strip().str.upper()

    entry_mask = (prev_hold == "OUT") & (curr_hold == "IN")
    bad_mask = entry_mask & (regime_series != target_name)

    violations: list[dict[str, Any]] = []

    date_col = _pick_first_existing_column(df, ["Date", "Datetime", "date", "datetime", "timestamp"])

    for ix in df.index[bad_mask].tolist()[:50]:
        item: dict[str, Any] = {
            "row_index": int(ix),
            "regime_found": str(regime_series.loc[ix]),
        }
        if date_col is not None:
            item["date"] = str(df.loc[ix, date_col])
        violations.append(item)

    return {
        "ok": len(violations) == 0,
        "reason": "" if len(violations) == 0 else "entry_outside_target_regime",
        "entry_count": int(entry_mask.sum()),
        "bad_entry_count": int(bad_mask.sum()),
        "target_regime_name": target_name,
        "signal_csv": str(signal_csv),
        "violations": violations,
    }

def _apply_gate_check_to_objective(
    obj: ObjectiveResult,
    gate_check: dict[str, Any],
    *,
    penalty_hard: float = 1_000_000_000.0,
) -> ObjectiveResult:
    if gate_check.get("ok", False):
        return obj

    reason = gate_check.get("reason") or "entry_outside_target_regime"
    penalty = max(float(obj.penalty), float(penalty_hard))

    return ObjectiveResult(
        alpha_vs_buyhold=obj.alpha_vs_buyhold,
        score=-float(penalty_hard),
        penalty=penalty,
        ok=False,
        reason=reason,
    )


def _read_final_composed_signal_csv(
    outdir: Path,
    selection_result: dict | None,
) -> Path | None:
    """
    Resolve the canonical SIGNAL csv of the final composed strategy.

    Priority:
    1) final_report/SIGNAL_*.csv
    2) last accepted forward-selection step under selection/eval/step_XX_<REGIME>
    """
    outdir = Path(outdir)

    # 1) Canonical final report location
    final_report_dir = outdir / "final_report"
    if final_report_dir.exists():
        candidates = [
            p for p in final_report_dir.rglob("SIGNAL_*.csv")
            if p.is_file() and not p.name.startswith("TRADE_FREQ_")
        ]
        if candidates:
            candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
            return candidates[0]

    # 2) Fallback: last accepted step in selection/eval
    if not selection_result:
        return None

    selected = selection_result.get("selected_regimes") or []
    if not selected:
        return None

    last_idx = len(selected) - 1
    last_candidate = selected[-1]
    step_name = f"step_{last_idx:02d}_{last_candidate}"

    step_dir = outdir / "selection" / "eval" / step_name
    if not step_dir.exists():
        return None

    candidates = [
        p for p in step_dir.rglob("SIGNAL_*.csv")
        if p.is_file() and not p.name.startswith("TRADE_FREQ_")
    ]
    if not candidates:
        return None

    candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]

def _read_final_composed_trade_rows(
    outdir: Path,
    selection_result: dict | None,
) -> list[dict[str, object]]:
    """
    Return vertical trade-detail rows for the final composed strategy:
    #trade, Profit del Trade, regime apertura, regime chiusura

    Notes:
    - Profit/Trade is read from the final SIGNAL csv.
    - Regime apertura / chiusura are mapped from dedicated columns if present.
    - If only a generic REGIME_L1_CODE column exists, it is used as fallback.
    """
    signal_csv = _read_final_composed_signal_csv(outdir, selection_result)
    if signal_csv is None:
        return []

    try:
        df = pd.read_csv(signal_csv, sep=";", engine="python")
    except Exception:
        return []

    if "Profit/Trade" not in df.columns:
        return []

    cols = {str(c).strip().lower(): c for c in df.columns}

    preferred_open = [
        "regime_open",
        "entry_regime",
        "REGIME_OPEN",
        "ENTRY_REGIME",
        "regime_at_entry",
        "REGIME_AT_ENTRY",
        "REGIME_L1_CODE_entry",
        "REGIME_L1_CODE_ENTRY",
    ]
    preferred_close = [
        "regime_close",
        "exit_regime",
        "REGIME_CLOSE",
        "EXIT_REGIME",
        "regime_at_exit",
        "REGIME_AT_EXIT",
        "REGIME_L1_CODE_exit",
        "REGIME_L1_CODE_EXIT",
    ]

    open_regime_col = next((cols[n] for n in preferred_open if n in cols), None)
    close_regime_col = next((cols[n] for n in preferred_close if n in cols), None)
    generic_regime_col = cols.get("regime_l1_code")

    out_rows: list[dict[str, object]] = []
    trade_no = 0

    for _, row in df.iterrows():
        profit = _to_float_maybe(row.get("Profit/Trade"))
        if profit is None:
            continue

        trade_no += 1

        regime_open = ""
        regime_close = ""

        if open_regime_col is not None:
            regime_open = str(row.get(open_regime_col) or "").strip()
        elif generic_regime_col is not None:
            regime_open = _decode_regime_code(row.get(generic_regime_col))

        if close_regime_col is not None:
            regime_close = str(row.get(close_regime_col) or "").strip()
        elif generic_regime_col is not None:
            regime_close = _decode_regime_code(row.get(generic_regime_col))

        out_rows.append(
            {
                "trade_no": trade_no,
                "profit_per_trade": profit,
                "regime_open": regime_open,
                "regime_close": regime_close,
            }
        )

    return out_rows



def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def default_outdir(base_config: Path) -> Path:
    stem = base_config.stem
    return Path("_data") / "strategy_tuning_runs" / f"{_ts()}_{stem}"
# ===============================
# EU number formatting helpers
# ===============================

_EU_DECIMAL_RE = re.compile(r"^-?\d+(\.\d+)?$")

def _fmt_eu_num(v: Any) -> Any:
    """
    Convert numbers to EU decimal comma representation.
    12.34 -> "12,34"
    """
    if v is None:
        return v

    if isinstance(v, bool):
        return v

    if isinstance(v, (int, float)):
        s = f"{v}"
        if "e" in s.lower():
            s = f"{v:.12f}".rstrip("0").rstrip(".")
        return s.replace(".", ",")

    if isinstance(v, str):
        s = v.strip()
        if _EU_DECIMAL_RE.match(s):
            return s.replace(".", ",")
        return v

    return v


def _normalize_candidate_values_cell(v: Any) -> Any:
    """
    candidate_values:
    - '|' -> ';'
    - decimal '.' -> ','
    """
    if v is None:
        return v

    s = str(v).strip()

    if not s:
        return v

    s = s.replace("|", ";")

    parts = [p.strip() for p in s.split(";")]
    out = []

    for p in parts:
        out.append(str(_fmt_eu_num(p)))

    return ";".join(out)


def _normalize_best_xlsx(best_path: Path) -> None:
    """
    Normalize TUNING.candidate_values in best.xlsx
    """
    wb = load_workbook(best_path)

    if "TUNING" not in wb.sheetnames:
        wb.close()
        return

    ws = wb["TUNING"]

    header = [c.value for c in ws[1]]

    try:
        idx = header.index("candidate_values") + 1
    except ValueError:
        wb.close()
        return

    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=idx)
        c.value = _normalize_candidate_values_cell(c.value)

    wb.save(best_path)
    wb.close()


def _read_sheet_as_df_preserve_order(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Read an Excel sheet preserving the column order as returned by pandas/openpyxl.
    Keep original values to make the export immediately usable for Excel copy/paste.
    """
    xlsx_path = Path(xlsx_path)
    return pd.read_excel(xlsx_path, sheet_name=sheet_name, engine="openpyxl")


def _write_tab_csv(df: pd.DataFrame, out_csv: Path) -> None:
    out_csv = Path(out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(out_csv, sep="\t", index=False, header=True, encoding="utf-8")


def _pick_existing_col_case_insensitive(df: pd.DataFrame, candidates: list[str]) -> str | None:
    lut = {str(c).strip().lower(): c for c in df.columns}
    for name in candidates:
        key = str(name).strip().lower()
        if key in lut:
            return lut[key]
    return None


def _filter_conditions_df_for_group(conditions_df: pd.DataFrame, target_group: str) -> pd.DataFrame:
    group_col = _pick_existing_col_case_insensitive(conditions_df, ["group"])
    if group_col is None:
        raise ValueError("CONDITIONS sheet missing 'group' column")

    target = str(target_group).strip().upper()
    mask = (
        conditions_df[group_col]
        .astype(str)
        .str.strip()
        .str.upper()
        == target
    )
    return conditions_df.loc[mask].copy()


def _extract_condition_ids_for_group(conditions_df: pd.DataFrame, target_group: str) -> list[str]:
    id_col = _pick_existing_col_case_insensitive(conditions_df, ["id"])
    if id_col is None:
        raise ValueError("CONDITIONS sheet missing 'id' column")

    block_df = _filter_conditions_df_for_group(conditions_df, target_group)

    out: list[str] = []
    for v in block_df[id_col].tolist():
        s = str(v).strip() if v is not None else ""
        if s:
            out.append(s)
    return out


def _filter_tuning_df_for_condition_ids(
    tuning_df: pd.DataFrame,
    condition_ids: list[str],
) -> pd.DataFrame:
    key_col = _pick_existing_col_case_insensitive(
        tuning_df,
        ["base_condition_id", "condition_id", "id"],
    )
    if key_col is None:
        raise ValueError(
            "TUNING sheet missing key column: expected one of "
            "['base_condition_id', 'condition_id', 'id']"
        )

    wanted = {str(x).strip() for x in condition_ids if str(x).strip()}
    mask = tuning_df[key_col].astype(str).str.strip().isin(wanted)
    return tuning_df.loc[mask].copy()


def _export_named_block_csvs(
    *,
    conditions_df: pd.DataFrame,
    tuning_df: pd.DataFrame,
    out_conditions_csv: Path,
    out_tuning_csv: Path,
) -> None:
    _write_tab_csv(conditions_df, out_conditions_csv)
    _write_tab_csv(tuning_df, out_tuning_csv)


def _export_group_block_csvs_from_workbook(
    *,
    source_xlsx: Path,
    base_tuning_xlsx: Path,
    target_group: str,
    out_dirs: list[Path],
    conditions_filename: str = "best_CONDITIONS.csv",
    tuning_filename: str = "best_TUNING.csv",
) -> None:
    """
    Export one group/block:
    - CONDITIONS from source_xlsx filtered by group
    - TUNING from base_tuning_xlsx filtered by base_condition_id/condition_id/id
      using the condition ids found in the selected CONDITIONS block
    """
    source_xlsx = Path(source_xlsx)
    base_tuning_xlsx = Path(base_tuning_xlsx)

    conditions_df = _read_sheet_as_df_preserve_order(source_xlsx, "CONDITIONS")
    block_conditions_df = _filter_conditions_df_for_group(conditions_df, target_group)
    condition_ids = _extract_condition_ids_for_group(conditions_df, target_group)

    try:
        tuning_df = _read_sheet_as_df_preserve_order(base_tuning_xlsx, "TUNING")
        block_tuning_df = _filter_tuning_df_for_condition_ids(tuning_df, condition_ids)
    except Exception:
        # Keep export robust even if TUNING is absent or structurally different
        block_tuning_df = pd.DataFrame()

    for out_dir in out_dirs:
        out_dir = Path(out_dir)
        _export_named_block_csvs(
            conditions_df=block_conditions_df,
            tuning_df=block_tuning_df,
            out_conditions_csv=out_dir / conditions_filename,
            out_tuning_csv=out_dir / tuning_filename,
        )


def _export_multi_group_block_csvs_from_workbook(
    *,
    source_xlsx: Path,
    base_tuning_xlsx: Path,
    target_groups: list[str],
    out_dir: Path,
    conditions_filename: str = "best_CONDITIONS.csv",
    tuning_filename: str = "best_TUNING.csv",
) -> None:
    """
    Export multiple selected groups from a composed workbook:
    - CONDITIONS from source_xlsx filtered by selected groups
    - TUNING from base_tuning_xlsx filtered by the union of involved condition ids
    """
    source_xlsx = Path(source_xlsx)
    base_tuning_xlsx = Path(base_tuning_xlsx)
    out_dir = Path(out_dir)

    conditions_df = _read_sheet_as_df_preserve_order(source_xlsx, "CONDITIONS")

    group_col = _pick_existing_col_case_insensitive(conditions_df, ["group"])
    id_col = _pick_existing_col_case_insensitive(conditions_df, ["id"])
    if group_col is None or id_col is None:
        raise ValueError("CONDITIONS sheet missing 'group' and/or 'id' column")

    wanted_groups = {
        str(g).strip().upper()
        for g in (target_groups or [])
        if str(g).strip()
    }

    cond_mask = (
        conditions_df[group_col]
        .astype(str)
        .str.strip()
        .str.upper()
        .isin(wanted_groups)
    )
    selected_conditions_df = conditions_df.loc[cond_mask].copy()

    condition_ids = [
        str(v).strip()
        for v in selected_conditions_df[id_col].tolist()
        if v is not None and str(v).strip()
    ]

    try:
        tuning_df = _read_sheet_as_df_preserve_order(base_tuning_xlsx, "TUNING")
        selected_tuning_df = _filter_tuning_df_for_condition_ids(tuning_df, condition_ids)
    except Exception:
        selected_tuning_df = pd.DataFrame()

    _export_named_block_csvs(
        conditions_df=selected_conditions_df,
        tuning_df=selected_tuning_df,
        out_conditions_csv=out_dir / conditions_filename,
        out_tuning_csv=out_dir / tuning_filename,
    )


def _resolve_best_composed_xlsx(outdir: Path, selection_result: dict | None) -> Path | None:
    """
    Resolve the final composed workbook path with conservative fallbacks.
    """
    outdir = Path(outdir)

    candidates: list[Path] = [
        outdir / "selection" / "best_composed.xlsx",
        outdir / "selection" / "baseline_composed.xlsx",
    ]

    if isinstance(selection_result, dict):
        for key in (
            "best_composed_xlsx",
            "composed_xlsx",
            "selected_config_xlsx",
            "best_config_xlsx",
        ):
            value = selection_result.get(key)
            if value:
                candidates.append(Path(value))

    for path in candidates:
        if path.exists() and path.is_file():
            return path

    return None



def _safe_group_dirname(group_name: str) -> str:
    s = str(group_name).strip()
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", s)
    return s or "UNKNOWN_GROUP"


def _read_conditions_rows(config_strategy: Path) -> Tuple[Any, List[str], Dict[str, int], List[Dict[str, Any]]]:
    p = Path(config_strategy)

    try:
        with open(p, "rb") as fh:
            sig = fh.read(8)
    except Exception:
        sig = None


    from openpyxl.utils.exceptions import InvalidFileException

    try:
        wb = load_workbook(p)
    except InvalidFileException as e:
        raise RuntimeError(
            f"load_workbook failed for config_strategy={p} "
            f"(suffix={p.suffix!r}, size={p.stat().st_size if p.exists() else None}, sig={sig!r})"
        ) from e

    if "CONDITIONS" not in wb.sheetnames:
        wb.close()
        raise ValueError("Missing CONDITIONS sheet in config_strategy")

    ws = wb["CONDITIONS"]
    header = [c.value for c in ws[1]]
    idx = {str(name): i + 1 for i, name in enumerate(header) if name is not None}

    required = ["id", "enabled", "scope", "side", "group"]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing CONDITIONS columns: {missing}")

    rows: List[Dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row: Dict[str, Any] = {}
        for col_name, col_idx in idx.items():
            row[col_name] = ws.cell(row=r, column=col_idx).value
        rows.append(row)

    return wb, header, idx, rows


def _is_truthy_enabled(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "on"}


def _normalize_scope(v: Any) -> str:
    return str(v).strip().upper() if v is not None else ""


def _normalize_side(v: Any) -> str:
    return str(v).strip().upper() if v is not None else ""


def _normalize_group(v: Any) -> str:
    return str(v).strip() if v is not None else ""



from pathlib import Path
from openpyxl import load_workbook


def _norm_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_upper(value) -> str:
    return _norm_text(value).upper()


def _is_enabled_cell(value) -> bool:
    if isinstance(value, bool):
        return value
    s = _norm_upper(value)
    return s in {"TRUE", "1", "YES", "Y", "ON"}

def detect_entry_groups(config_strategy) -> list[str]:
    """
    Rileva i group che hanno almeno una riga ENTRY abilitata
    nel foglio CONDITIONS del workbook config_strategy.

    Nota architetturale:
    G_TREND non è un gruppo operativo del framework.
    I soli regimi trend validi per strategy design, autotuning e selection
    sono G_TREND_UP e G_TREND_DOWN.
    """
    path = Path(config_strategy)
    wb = load_workbook(path, data_only=False)

    if "CONDITIONS" not in wb.sheetnames:
        return []

    ws = wb["CONDITIONS"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_norm_text(x) for x in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}

    required = {"enabled", "scope", "group"}
    if not required.issubset(idx.keys()):
        return []

    CANONICAL_REGIME_GROUPS = {
        "G_RANGE",
        "G_LATERAL",
        "G_VOLATILE",
        "G_TREND_UP",
        "G_TREND_DOWN",
    }

    LEGACY_EXCLUDED_REGIME_GROUPS = {
        "G_TREND",
    }

    out = []
    seen = set()
    ignored_legacy = set()

    for row in rows[1:]:
        group = _norm_upper(row[idx["group"]])
        scope = _norm_upper(row[idx["scope"]])
        enabled = row[idx["enabled"]]

        if not group:
            continue
        if scope != "ENTRY":
            continue
        if not _is_enabled_cell(enabled):
            continue

        if group in LEGACY_EXCLUDED_REGIME_GROUPS:
            ignored_legacy.add(group)
            continue

        if group not in CANONICAL_REGIME_GROUPS:
            continue

        if group not in seen:
            seen.add(group)
            out.append(group)

    if ignored_legacy:
        print(
            "[INFO] Ignoro gruppi legacy non operativi: "
            + ", ".join(sorted(ignored_legacy))
        )

    return out

def _debug_dump_entry_rows(config_strategy) -> list[dict[str, Any]]:
    path = Path(config_strategy)
    wb = load_workbook(path, data_only=False)

    if "CONDITIONS" not in wb.sheetnames:
        return []

    ws = wb["CONDITIONS"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_norm_text(x) for x in rows[0]]
    idx = {name: i for i, name in enumerate(headers)}

    required = {"id", "enabled", "scope", "side", "group"}
    if not required.issubset(idx.keys()):
        return []

    out = []
    for row in rows[1:]:
        item = {
            "id": _norm_text(row[idx["id"]]),
            "enabled": row[idx["enabled"]],
            "scope": _norm_text(row[idx["scope"]]),
            "side": _norm_text(row[idx["side"]]),
            "group": _norm_text(row[idx["group"]]),
        }
        if (
            item["group"]
            and _norm_upper(item["scope"]) == "ENTRY"
            and _is_enabled_cell(item["enabled"])
        ):
            out.append(item)
    return out


def _print_debug_entry_rows(config_strategy) -> None:
    try:
        rows = _debug_dump_entry_rows(config_strategy)
        print(f"[DBG] active ENTRY rows from CONDITIONS = {len(rows)}")
        for item in rows:
            print(
                "[DBG] ENTRY "
                f"id={item['id']} "
                f"enabled={item['enabled']} "
                f"scope={item['scope']} "
                f"side={item['side']} "
                f"group={item['group']}"
            )
    except Exception as exc:
        print(f"[DBG] unable to dump entry rows: {exc}")


def _decision_hint(
    *,
    trade_count: Any,
    alpha_vs_buyhold: Any,
    n_min_trades: int,
) -> str:
    try:
        tc = 0 if trade_count is None else int(float(str(trade_count).replace(",", ".")))
    except Exception:
        tc = 0

    try:
        alpha = None if alpha_vs_buyhold is None else float(str(alpha_vs_buyhold).replace(",", "."))
    except Exception:
        alpha = None

    if tc <= 0:
        return "DROP"

    if tc < int(n_min_trades):
        return "REVIEW"

    if alpha is None:
        return "REVIEW"

    if alpha > 0:
        return "KEEP"

    return "DROP"

def _write_masked_conditions_config(
    *,
    base_config_path: Path,
    out_config_path: Path,
    active_entry_group: str,
) -> None:
    """
    Deprecated shim for backward compatibility.

    v3.0 semantics:
    - full dataset
    - temporary config with only target group enabled
    - no regime sub-dataset
    """
    build_regime_wise_config_v3(
        base_xlsx_path=base_config_path,
        out_xlsx_path=out_config_path,
        target_group=active_entry_group,
    )



def _merge_forward_conditions_from_blocks(
    *,
    selected_blocks: List[RegimeBlock],
    out_config_path: Path,
) -> None:
    """
    Merge CONDITIONS rows from selected per-regime best.xlsx files into an already
    composed config.

    Expected workflow:
    - out_config_path already exists and contains the baseline/composed CONDITIONS
      produced from the external template
    - for each selected block, read its CONDITIONS sheet
    - copy row-level editable fields for rows belonging to block.regime
    - match rows primarily by 'id'
    - preserve rows from other groups as already written in out_config_path

    Why this is needed:
    forward selection cannot rely on TUNING-only merge because many strategy mutations
    live directly in CONDITIONS (e.g. rhs_value / operator / logic / enabled).
    """
    if not selected_blocks:
        return

    composed_wb = load_workbook(out_config_path)
    if "CONDITIONS" not in composed_wb.sheetnames:
        composed_wb.close()
        raise ValueError("Missing CONDITIONS sheet in composed config")

    composed_ws = composed_wb["CONDITIONS"]
    composed_header = [c.value for c in composed_ws[1]]
    composed_idx = {str(name): i + 1 for i, name in enumerate(composed_header) if name is not None}

    required = ["id", "group"]
    missing = [c for c in required if c not in composed_idx]
    if missing:
        composed_wb.close()
        raise ValueError(f"Missing CONDITIONS columns in composed config: {missing}")

    editable_cols = ["enabled", "rhs_value", "shift"]
    editable_cols = [c for c in editable_cols if c in composed_idx]

    composed_row_by_id: Dict[str, int] = {}
    for r in range(2, composed_ws.max_row + 1):
        row_id = composed_ws.cell(r, composed_idx["id"]).value
        if row_id is None:
            continue
        composed_row_by_id[str(row_id)] = r

    for block in selected_blocks:
        block_wb = load_workbook(block.best_xlsx, data_only=False)
        try:
            if "CONDITIONS" not in block_wb.sheetnames:
                raise ValueError(f"Missing CONDITIONS sheet in {block.best_xlsx}")

            block_ws = block_wb["CONDITIONS"]
            block_header = [c.value for c in block_ws[1]]
            block_idx = {str(name): i + 1 for i, name in enumerate(block_header) if name is not None}

            block_required = ["id", "group"]
            block_missing = [c for c in block_required if c not in block_idx]
            if block_missing:
                raise ValueError(
                    f"Missing CONDITIONS columns in {block.best_xlsx.name}: {block_missing}"
                )

            target_group = str(getattr(block, "regime", "") or "").strip().upper()
            if not target_group:
                continue

            for r in range(2, block_ws.max_row + 1):
                row_id = block_ws.cell(r, block_idx["id"]).value
                row_group = block_ws.cell(r, block_idx["group"]).value

                if row_id is None:
                    continue

                row_group_norm = str(row_group).strip().upper() if row_group is not None else ""
                if row_group_norm != target_group:
                    continue

                row_id_str = str(row_id)
                composed_r = composed_row_by_id.get(row_id_str)
                if composed_r is None:
                    continue

                for col in editable_cols:
                    if col not in block_idx:
                        continue
                    composed_ws.cell(composed_r, composed_idx[col]).value = block_ws.cell(
                        r, block_idx[col]
                    ).value
        finally:
            block_wb.close()

    out_config_path.parent.mkdir(parents=True, exist_ok=True)
    composed_wb.save(out_config_path)
    composed_wb.close()


def _write_forward_composed_conditions_config(
    *,
    base_config_path: Path,
    out_config_path: Path,
    active_groups: List[str],
) -> None:
    """
    Build a composed strategy config from the external base template.

    Semantics:
    - keep global groups (e.g. G_ANY) always enabled
    - enable rows whose CONDITIONS.group belongs to active_groups
    - disable all other rows
    """
    wb = load_workbook(base_config_path)
    if "CONDITIONS" not in wb.sheetnames:
        wb.close()
        raise ValueError("Missing CONDITIONS sheet in config_strategy")

    ws = wb["CONDITIONS"]
    header = [c.value for c in ws[1]]
    idx = {str(name): i + 1 for i, name in enumerate(header) if name is not None}

    required = ["enabled", "group"]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing CONDITIONS columns: {missing}")

    normalized_active = {
        _normalize_group(g).upper()
        for g in (active_groups or [])
        if _normalize_group(g)
    }

    for r in range(2, ws.max_row + 1):
        group_raw = ws.cell(r, idx["group"]).value
        group_norm = _normalize_group(group_raw)
        group_upper = group_norm.upper()

        keep_enabled = False
        if _is_global_group(group_norm):
            keep_enabled = True
        elif group_upper in normalized_active:
            keep_enabled = True

        ws.cell(r, idx["enabled"]).value = keep_enabled

    out_config_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_config_path)
    wb.close()
def _merge_forward_tuning_from_blocks(
    *,
    base_config_path: Path,
    selected_blocks: List[RegimeBlock],
    out_config_path: Path,
) -> None:
    """
    Merge TUNING sheet values from selected per-regime best.xlsx files into a composed config.

    Strategy:
    - start from TUNING of base_config_path
    - for each selected block, read TUNING from block.best_xlsx
    - merge rows by parameter id (column 'param_name' if present, else 'id')
    - for overlapping parameter ids, later blocks override earlier ones
    - write merged TUNING into out_config_path, preserving all non-TUNING sheets
    """
    if not selected_blocks:
        return

    base_df = read_tuning_sheet(base_config_path).copy()

    key_col = None
    for c in ("param_name", "id"):
        if c in base_df.columns:
            key_col = c
            break
    if key_col is None:
        raise ValueError("TUNING sheet missing key column 'param_name' or 'id'")

    merged_df = base_df.copy()
    merged_df[key_col] = merged_df[key_col].astype(str)

    for block in selected_blocks:
        block_df = read_tuning_sheet(block.best_xlsx).copy()
        if key_col not in block_df.columns:
            raise ValueError(
                f"TUNING sheet in {block.best_xlsx.name} missing key column '{key_col}'"
            )

        block_df[key_col] = block_df[key_col].astype(str)

        # Use block rows as authoritative for matching parameter ids
        block_map = block_df.set_index(key_col)
        merged_df = merged_df.set_index(key_col)

        common_keys = [k for k in merged_df.index if k in block_map.index]
        for k in common_keys:
            for col in merged_df.columns:
                if col in block_map.columns:
                    merged_df.at[k, col] = block_map.at[k, col]

        merged_df = merged_df.reset_index()

    write_tuning_sheet(
        base_xlsx_path=out_config_path,
        tuning_df=merged_df,
        out_xlsx_path=out_config_path,
    )
CANONICAL_REGIME_GROUPS = {
    "G_RANGE",
    "G_LATERAL",
    "G_VOLATILE",
    "G_TREND_UP",
    "G_TREND_DOWN",
}

LEGACY_EXCLUDED_REGIME_GROUPS = {
    "G_TREND",
}

def _is_canonical_regime_group(group_name: str) -> bool:
    return str(group_name).strip().upper() in CANONICAL_REGIME_GROUPS

def _is_legacy_excluded_regime_group(group_name: str) -> bool:
    return str(group_name).strip().upper() in LEGACY_EXCLUDED_REGIME_GROUPS

def _filter_operational_regime_groups(groups: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    for g in groups:
        gs = str(g).strip().upper()
        if not gs:
            continue
        if _is_legacy_excluded_regime_group(gs):
            continue
        if not _is_canonical_regime_group(gs):
            continue
        if gs in seen:
            continue
        seen.add(gs)
        out.append(gs)

    return out

def _is_global_group(group_value: Any) -> bool:
    """
    Global/non-regime groups that should remain enabled even when the
    forward-selection baseline represents S = {}.

    Convention:
    - keep G_ANY active
    - empty group is treated as non-global here (disabled by default)
    """
    g = _normalize_group(group_value).upper()
    return g in {"G_ANY"}


def _read_trade_profit_list_from_best_signal(best_xlsx_path) -> list[float]:
    """
    Backward-compatible shim.

    IMPORTANT:
    - Use the same trade extraction semantics as the canonical standalone summary:
      every non-null Profit/Trade value from the resolved SIGNAL_*.csv is a closed trade.
    - Do NOT drop zero-profit trades.
    - Do NOT use ad-hoc CSV parsing that may disagree with evaluator.py.
    """
    return _read_trade_profit_list_from_best_xlsx(best_xlsx_path)

def _write_forward_baseline_config(
    *,
    base_config_path: Path,
    out_config_path: Path,
) -> None:
    """
    Build the true forward-selection baseline S = {}.

    Semantics:
    - disable every regime-specific CONDITION row
    - keep enabled only global rows (currently group == G_ANY)
    - preserve all non-CONDITIONS sheets as-is
    """
    wb = load_workbook(base_config_path)
    if "CONDITIONS" not in wb.sheetnames:
        wb.close()
        raise ValueError("Missing CONDITIONS sheet in config_strategy")

    ws = wb["CONDITIONS"]
    header = [c.value for c in ws[1]]
    idx = {str(name): i + 1 for i, name in enumerate(header) if name is not None}

    required = ["enabled", "group"]
    missing = [c for c in required if c not in idx]
    if missing:
        wb.close()
        raise ValueError(f"Missing CONDITIONS columns: {missing}")

    for r in range(2, ws.max_row + 1):
        group = ws.cell(r, idx["group"]).value
        ws.cell(r, idx["enabled"]).value = _is_global_group(group)

    out_config_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_config_path)
    wb.close()


def run_autotune_v1(
        *,
        input_csv: Path,
        config_strategy: Path,
        timeframe: str,
        trials: int = 50,
        seed: int = 42,
        outdir: Optional[Path] = None,
        n_min_trades: int = 5,
        timeout_sec: int = 300,
        active_group: Optional[str] = None,
        trial_label: Optional[str] = None,
) -> Path:
    input_csv = Path(input_csv)
    config_strategy = Path(config_strategy)
    if outdir is None:
        outdir = default_outdir(config_strategy)

    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    # dirs
    trials_dir = outdir / "trials"
    eval_dir_root = outdir / "eval"
    trials_dir.mkdir(exist_ok=True)
    eval_dir_root.mkdir(exist_ok=True)

    runspec = RunSpec(
        input_csv=str(input_csv),
        config_strategy=str(config_strategy),
        timeframe=timeframe,
        trials=int(trials),
        seed=int(seed),
        outdir=str(outdir),
        n_min_trades=int(n_min_trades),
        active_group=active_group,
    )

    (outdir / "runspec.json").write_text(json.dumps(asdict(runspec), indent=2), encoding="utf-8")

    # build space + sampler
    space = build_space_from_tuning(config_strategy, active_group=active_group)

    if not space.params:
        print(
            f"[WARN] No tunable parameters for group={active_group}. "
            "Skipping tuning and evaluating baseline regime block. "
            "This regime can still be proposed during forward selection "
            "if its baseline block improves the composed strategy."
        )

        evaluator = StrategyEvaluator()

        eval_dir = outdir / "eval_baseline"
        eval_dir.mkdir(parents=True, exist_ok=True)

        if active_group:
            ev = evaluator.evaluate_regime_wise_v3(
                input_csv=input_csv,
                config_strategy=config_strategy,
                timeframe=timeframe,
                outdir=eval_dir,
                timeout_sec=timeout_sec,
                target_regime_name=_group_to_regime_name(active_group),
            )
        else:
            ev = evaluator.evaluate(
                input_csv=input_csv,
                config_strategy=config_strategy,
                timeframe=timeframe,
                outdir=eval_dir,
                timeout_sec=timeout_sec,
            )

        obj_spec = ObjectiveSpec(
            n_min_trades=n_min_trades,
            drawdown_weight=0.5,
        )

        obj = compute_objective(ev.metrics, obj_spec)


        if active_group:
            signal_csv_str = (getattr(ev.metrics, "extras", {}) or {}).get("signal_csv")
            if signal_csv_str:
                gate_check = _verify_entries_only_on_target_regime(
                    signal_csv=Path(signal_csv_str),
                    target_regime_name=_group_to_regime_name(active_group),
                )
            else:
                gate_check = {
                    "ok": False,
                    "reason": "missing_signal_csv",
                    "entry_count": 0,
                    "bad_entry_count": 0,
                    "violations": [],
                }

            extras = dict(getattr(ev.metrics, "extras", {}) or {})
            extras["regime_gate_check"] = gate_check
            extras["regime_eval_mode"] = "v3_full_dataset_target_entries_only"
            extras["target_group"] = active_group
            ev.metrics.extras = extras

            obj = _apply_gate_check_to_objective(obj, gate_check)

        trials_csv = outdir / "trials.csv"

        fieldnames = [
            "trial_id",
            "score",
            "alpha_vs_buyhold",
            "penalty",
            "ok",
            "reason",
            "profit",
            "profit_per_trade",
            "trade_count",
            "max_drawdown",
            "buy_hold_profit",
            "trial_config_xlsx",
            "eval_dir",
            "evaluation_semantics",
            "gate_check_ok",
            "bad_entry_count",
        ]

        with trials_csv.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
            w.writeheader()

            gate_check = (getattr(ev.metrics, "extras", {}) or {}).get("regime_gate_check", {}) or {}

            row = {
                "trial_id": 0,
                "score": obj.score,
                "alpha_vs_buyhold": obj.alpha_vs_buyhold,
                "penalty": obj.penalty,
                "ok": obj.ok,
                "reason": obj.reason,
                "profit": ev.metrics.profit,
                "profit_per_trade": ev.metrics.profit_per_trade,
                "trade_count": ev.metrics.trade_count,
                "max_drawdown": ev.metrics.max_drawdown,
                "buy_hold_profit": ev.metrics.buy_hold_profit,
                "trial_config_xlsx": str(config_strategy),
                "eval_dir": str(eval_dir),
                "evaluation_semantics": (getattr(ev.metrics, "extras", {}) or {}).get("regime_eval_mode"),
                "gate_check_ok": gate_check.get("ok"),
                "bad_entry_count": gate_check.get("bad_entry_count"),
            }

            w.writerow({k: _fmt_eu_num(v) for k, v in row.items()})

        best_path = outdir / "best.xlsx"
        best_path.write_bytes(config_strategy.read_bytes())

        try:
            _normalize_best_xlsx(best_path)
        except Exception:
            pass


        return outdir

    sampler = RandomSearchSampler(seed=seed)
    obj_spec = ObjectiveSpec(
        n_min_trades=n_min_trades,
        drawdown_weight=0.5,
    )

    evaluator = StrategyEvaluator()

    # open trials.csv
    trials_csv = outdir / "trials.csv"
    fieldnames = [
                     "trial_id",
                     "score",
                     "alpha_vs_buyhold",
                     "penalty",
                     "ok",
                     "reason",
                     "profit",
                     "profit_per_trade",
                     "trade_count",
                     "max_drawdown",
                     "buy_hold_profit",
                     "trial_config_xlsx",
                     "eval_dir",
                     "evaluation_semantics",
                     "gate_check_ok",
                     "bad_entry_count",
    ] + [p.name for p in space.params]

    best_score = float("-inf")
    best_trial_xlsx: Optional[Path] = None

    with trials_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        w.writeheader()

        for sample in sampler.iter_trials(space, trials):
            tid = sample.trial_id

            trial_config_xlsx = trials_dir / f"trial_{tid:04d}.xlsx"
            eval_dir = eval_dir_root / f"trial_{tid:04d}"
            eval_dir.mkdir(parents=True, exist_ok=True)

            # write trial config (ONLY TUNING)
            write_trial_config(
                base_config_path=config_strategy,
                out_config_path=trial_config_xlsx,
                params=sample.params,
            )

            # evaluate
            if active_group:
                ev = evaluator.evaluate_regime_wise_v3(
                    input_csv=input_csv,
                    config_strategy=trial_config_xlsx,
                    timeframe=timeframe,
                    outdir=eval_dir,
                    timeout_sec=timeout_sec,
                    target_regime_name=_group_to_regime_name(active_group),
                )
            else:
                ev = evaluator.evaluate(
                    input_csv=input_csv,
                    config_strategy=trial_config_xlsx,
                    timeframe=timeframe,
                    outdir=eval_dir,
                    timeout_sec=timeout_sec,
                )

            # regime-wise v3 gate-check:
            # enforce only when active_group is set (single-regime tuning)
            gate_check = {}
            if active_group:
                signal_csv_str = (getattr(ev.metrics, "extras", {}) or {}).get("signal_csv")
                if signal_csv_str:
                    gate_check = _verify_entries_only_on_target_regime(
                        signal_csv=Path(signal_csv_str),
                        target_regime_name=_group_to_regime_name(active_group),
                    )
                else:
                    gate_check = {
                        "ok": False,
                        "reason": "missing_signal_csv",
                        "entry_count": 0,
                        "bad_entry_count": 0,
                        "violations": [],
                    }

                extras = dict(getattr(ev.metrics, "extras", {}) or {})
                extras["regime_gate_check"] = gate_check
                extras["regime_eval_mode"] = "v3_full_dataset_target_entries_only"
                extras["target_group"] = active_group
                ev.metrics.extras = extras

            # objective
            obj = compute_objective(ev.metrics, obj_spec)

            if active_group and gate_check:
                obj = _apply_gate_check_to_objective(obj, gate_check)

            display_best = max(best_score, obj.score) if best_score != float("-inf") else obj.score
            label = f"[{trial_label}] " if trial_label else ""
            print(
                f"\r{label}Trial {tid}/{trials}  score={obj.score:.2f}  best={display_best:.2f}",
                end="",
                flush=True,
            )

            gate_check = (getattr(ev.metrics, "extras", {}) or {}).get("regime_gate_check", {}) or {}

            row: Dict[str, Any] = {
                "trial_id": tid,
                "score": obj.score,
                "alpha_vs_buyhold": obj.alpha_vs_buyhold,
                "penalty": obj.penalty,
                "ok": obj.ok,
                "reason": obj.reason if ev.ok else (ev.error or obj.reason),
                "profit": ev.metrics.profit,
                "profit_per_trade": ev.metrics.profit_per_trade,
                "trade_count": ev.metrics.trade_count,
                "max_drawdown": ev.metrics.max_drawdown,
                "buy_hold_profit": ev.metrics.buy_hold_profit,
                "trial_config_xlsx": str(trial_config_xlsx),
                "eval_dir": str(eval_dir),
                "evaluation_semantics": (getattr(ev.metrics, "extras", {}) or {}).get("regime_eval_mode"),
                "gate_check_ok": gate_check.get("ok"),
                "bad_entry_count": gate_check.get("bad_entry_count"),
            }
            # params
            for k, v in sample.params.items():
                row[k] = v

            row_eu = {k: _fmt_eu_num(v) for k, v in row.items()}
            w.writerow(row_eu)

            if obj.score > best_score:
                best_score = obj.score
                best_trial_xlsx = trial_config_xlsx
    print()

    # save best.xlsx
    if best_trial_xlsx is not None:
        best_path = outdir / "best.xlsx"
        best_path.write_bytes(best_trial_xlsx.read_bytes())

        # normalize candidate_values and EU decimals
        try:
            _normalize_best_xlsx(best_path)
        except Exception:
            pass

    return outdir

def run_autotune_regimes_v1(
    *,
    input_csv: Path,
    config_strategy: Path,
    timeframe: str,
    regimes: Optional[Iterable[str]] = None,
    trials: int = 50,
    seed: int = 42,
    train_ratio: float = 0.7,
    outdir: Optional[Path] = None,
    n_min_trades: int = 5,
    timeout_sec: int = 300,
    interactive_selection: bool = False,

) -> Path:
    input_csv = Path(input_csv)
    config_strategy = Path(config_strategy)

    if outdir is None:
        outdir = default_outdir(config_strategy)

    outdir = Path(outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    per_regime_dir = outdir / "per_regime"
    selection_dir = outdir / "selection"
    final_report_dir = outdir / "final_report"
    per_regime_dir.mkdir(exist_ok=True)
    selection_dir.mkdir(exist_ok=True)
    final_report_dir.mkdir(exist_ok=True)

    detected_groups = detect_entry_groups(config_strategy)
    detected_groups = _filter_operational_regime_groups(detected_groups)
    print(f"[DBG] detect_entry_groups config_strategy = {Path(config_strategy).resolve()}")
    print(f"[DBG] detected_groups = {detected_groups}")

    _print_debug_entry_rows(config_strategy)


    # v3.0 semantics:
    # - detected_groups are derived only from config_strategy
    # - CLI --regimes is only a batch/orchestration filter
    if regimes:
        requested = [str(x).strip() for x in regimes if str(x).strip()]
        target_groups = [g for g in detected_groups if g in requested]
    else:
        target_groups = list(detected_groups)

        print(f"[DBG] requested_regimes = {list(regimes) if regimes else []}")
        print(f"[DBG] target_groups = {target_groups}")

    if regimes and not target_groups:
        requested_set = {str(x).strip() for x in regimes if str(x).strip()}
        detected_set = set(detected_groups)
        raise ValueError(
            "No target groups matched the current config_strategy. "
            f"requested={sorted(requested_set)} detected={sorted(detected_set)} "
            f"config_strategy={Path(config_strategy).resolve()}"
        )

    runspec = {
        "input_csv": str(input_csv),
        "config_strategy": str(config_strategy),
        "timeframe": timeframe,
        "trials_per_regime": int(trials),
        "seed": int(seed),
        "train_ratio": float(train_ratio),
        "outdir": str(outdir),
        "n_min_trades": int(n_min_trades),
        "requested_groups": [str(x).strip() for x in regimes if str(x).strip()] if regimes else [],
        "detected_entry_groups": detected_groups,
        "target_groups": target_groups,
    }

    (outdir / "runspec_regimes.json").write_text(
        json.dumps(runspec, indent=2),
        encoding="utf-8",
    )

    summary_rows: List[Dict[str, Any]] = []
    tuned_blocks: Dict[str, RegimeBlock] = {}

    for i, group_name in enumerate(target_groups, start=1):
        safe_group = _safe_group_dirname(group_name)
        regime_dir = per_regime_dir / f"regime_{safe_group}"
        regime_dir.mkdir(parents=True, exist_ok=True)

        masked_base = regime_dir / "regime_wise_v3_base.xlsx"
        build_regime_wise_config_v3(
            base_xlsx_path=config_strategy,
            out_xlsx_path=masked_base,
            target_group=group_name,
        )

        regime_seed = int(seed) + i - 1
        run_outdir = run_autotune_v1(
            input_csv=input_csv,
            config_strategy=masked_base,
            timeframe=timeframe,
            trials=int(trials),
            seed=regime_seed,
            outdir=regime_dir,
            n_min_trades=int(n_min_trades),
            timeout_sec=int(timeout_sec),
            active_group=group_name,
            trial_label=group_name,
        )

        trials_csv = run_outdir / "trials.csv"
        best_xlsx = run_outdir / "best.xlsx"

        best_profit = None
        best_trades = None
        best_alpha = None
        best_dd = None
        best_score = None

        # Source of truth aligned with trials.csv best row
        best_trial_id = None
        best_trial_config_xlsx = None
        selected_best_xlsx = best_xlsx
        selected_best_eval_dir = None

        df_sorted = pd.DataFrame()

        try:
            df = pd.read_csv(trials_csv, sep=";")

            if len(df) > 0:
                # Convert EU-decimal numeric columns before sorting/selecting best
                for col in [
                    "score",
                    "profit",
                    "trade_count",
                    "alpha_vs_buyhold",
                    "max_drawdown",
                    "buy_hold_profit",
                    "profit_per_trade",
                    "penalty",
                ]:
                    if col in df.columns:
                        df[col] = pd.to_numeric(
                            df[col].astype(str).str.replace(",", ".", regex=False),
                            errors="coerce",
                        )

                df_sorted = df.sort_values("score", ascending=False, na_position="last")
                best = df_sorted.iloc[0]

                best_profit = best.get("profit")
                best_trades = best.get("trade_count")
                best_alpha = best.get("alpha_vs_buyhold")
                best_dd = best.get("max_drawdown")
                best_score = best.get("score")

                best_trial_id = best.get("trial_id")
                best_trial_config_raw = best.get("trial_config_xlsx")
                best_eval_dir_raw = best.get("eval_dir")

                if best_eval_dir_raw is not None and str(best_eval_dir_raw).strip():
                    try:
                        selected_best_eval_dir = Path(str(best_eval_dir_raw))
                    except Exception:
                        selected_best_eval_dir = None

                if best_trial_config_raw is not None:
                    try:
                        best_trial_config_xlsx = Path(str(best_trial_config_raw))
                    except Exception:
                        best_trial_config_xlsx = None

                if best_trial_config_xlsx is not None and best_trial_config_xlsx.exists():
                    selected_best_xlsx = best_trial_config_xlsx
                else:
                    selected_best_xlsx = best_xlsx

                if (
                    best_trial_config_xlsx is not None
                    and best_trial_config_xlsx.exists()
                    and best_xlsx.exists()
                    and best_trial_config_xlsx.resolve() != best_xlsx.resolve()
                ):
                    print(
                        f"[WARN] regime={group_name} trials.csv best differs from run best.xlsx: "
                        f"trial_config_xlsx={best_trial_config_xlsx} best_xlsx={best_xlsx}"
                    )

                print(
                    f"[DBG] regime={group_name} "
                    f"best_trial_id={best_trial_id} "
                    f"best_score={best_score} "
                    f"best_trial_config_xlsx={best_trial_config_xlsx} "
                    f"selected_best_xlsx={selected_best_xlsx}"
                )

        except Exception as exc:
            print(f"[WARN] Could not read/sort trials for regime={group_name}: {exc}")

        try:
            _export_group_block_csvs_from_workbook(
                source_xlsx=selected_best_xlsx,
                base_tuning_xlsx=config_strategy,
                target_group=group_name,
                out_dirs=[regime_dir],
                conditions_filename="best_CONDITIONS.csv",
                tuning_filename="best_TUNING.csv",
            )

            _export_group_block_csvs_from_workbook(
                source_xlsx=selected_best_xlsx,
                base_tuning_xlsx=config_strategy,
                target_group=group_name,
                out_dirs=[final_report_dir],
                conditions_filename=f"best_{group_name}_CONDITIONS.csv",
                tuning_filename=f"best_{group_name}_TUNING.csv",
            )
        except Exception as exc:
            print(f"[WARN] regime={group_name} export best csv failed: {exc}")

        decision_hint = _decision_hint(
            trade_count=best_trades,
            alpha_vs_buyhold=best_alpha,
            n_min_trades=int(n_min_trades),
        )

        # Candidate block for forward selection/composition
        try:
            if selected_best_xlsx.exists():
                tuned_blocks[group_name] = RegimeBlock(
                    regime=group_name,
                    best_xlsx=selected_best_xlsx,
                    profit_trades=float(best_profit) if best_profit is not None and pd.notna(best_profit) else float("nan"),
                    profit_buyhold=(
                        float(best_profit) - float(best_alpha)
                        if best_profit is not None and best_alpha is not None and pd.notna(best_profit) and pd.notna(best_alpha)
                        else float("nan")
                    ),
                    alpha_regime=float(best_alpha) if best_alpha is not None and pd.notna(best_alpha) else float("nan"),
                    n_trades=int(best_trades) if best_trades is not None and pd.notna(best_trades) else 0,
                    max_dd=float(best_dd) if best_dd is not None and pd.notna(best_dd) else float("nan"),
                )
        except Exception:
            # Keep per-regime autotune robust even if selection metadata cannot be built
            pass

        best_profit_num = _to_float_maybe(best_profit)
        best_alpha_num = _to_float_maybe(best_alpha)
        best_trades_num = _to_float_maybe(best_trades)

        gate_check_ok = None
        bad_entry_count = None
        try:
            if len(df_sorted) > 0:
                best_gate = df_sorted.iloc[0]
                if "gate_check_ok" in best_gate.index:
                    gate_check_ok = best_gate.get("gate_check_ok")
                if "bad_entry_count" in best_gate.index:
                    bad_entry_count = best_gate.get("bad_entry_count")
        except Exception:
            pass

        summary_rows.append(
            {
                "group": group_name,
                "profit": best_profit_num if best_profit_num is not None else best_profit,
                "trade_count": int(best_trades_num) if best_trades_num is not None else best_trades,
                "profit_per_trade": _safe_profit_per_trade(best_profit_num, best_trades_num),
                "buy_hold_profit": (
                    best_profit_num - best_alpha_num
                    if best_profit_num is not None and best_alpha_num is not None
                    else None
                ),
                "alpha_vs_buyhold": best_alpha_num if best_alpha_num is not None else best_alpha,
                "max_drawdown": _to_float_maybe(best_dd) if _to_float_maybe(best_dd) is not None else best_dd,
                "score": _to_float_maybe(best_score) if _to_float_maybe(best_score) is not None else best_score,
                "standalone_hint": decision_hint,
                "best_xlsx": str(selected_best_xlsx),
                "eval_dir": str(selected_best_eval_dir) if selected_best_eval_dir else "",
                "gate_check_ok": gate_check_ok,
                "bad_entry_count": bad_entry_count,
            }
        )

    # ------------------------------------------------------------
    # Regime forward selection / composition
    # ------------------------------------------------------------
    selection_result = None
    try:
        candidate_regimes = [g for g in target_groups if g in tuned_blocks]

        # v2.1:
        # no hard-coded restriction of candidate regimes here.
        # Candidate screening / promotion / keep-for-joint-opt / hard-drop
        # must be decided by forward selection logic, not by an ad-hoc local filter.
        candidate_regimes = list(candidate_regimes)

        if candidate_regimes:
            fs_spec = ForwardSelectionSpec(
                candidate_regimes=candidate_regimes,
                eps_alpha=0.0,
                dd_tolerance_ratio=0.25,
                n_trades_min=int(n_min_trades),
                interactive=bool(interactive_selection),
                auto_accept=False,
                enforce_constraints=True,
                max_regimes=len(candidate_regimes),
            )

            def _forward_evaluator(
                input_csv: Path,
                config_strategy: Path,
                timeframe: str,
                outdir: Path,
            ) -> EvalResult:
                evaluator = StrategyEvaluator()
                ev = evaluator.evaluate(
                    input_csv=input_csv,
                    config_strategy=config_strategy,
                    timeframe=timeframe,
                    outdir=outdir,
                    timeout_sec=int(timeout_sec),
                )

                obj_spec = ObjectiveSpec(
                    n_min_trades=int(n_min_trades),
                    drawdown_weight=0.5,
                )
                obj = compute_objective(ev.metrics, obj_spec)

                alpha_val = (
                    float(ev.metrics.alpha_vs_buyhold)
                    if ev.metrics.alpha_vs_buyhold is not None
                    else float("nan")
                )
                trades_val = (
                    int(ev.metrics.trade_count)
                    if ev.metrics.trade_count is not None
                    else 0
                )
                dd_val = (
                    float(ev.metrics.max_drawdown)
                    if ev.metrics.max_drawdown is not None
                    else float("nan")
                )
                buy_hold_val = (
                    float(ev.metrics.buy_hold_profit)
                    if getattr(ev.metrics, "buy_hold_profit", None) is not None
                    else float("nan")
                )
                profit_val = (
                    float(ev.metrics.profit)
                    if getattr(ev.metrics, "profit", None) is not None
                    else float("nan")
                )
                penalty_val = (
                    float(obj.penalty)
                    if getattr(obj, "penalty", None) is not None
                    else 0.0
                )

                return EvalResult(
                    ok=bool(ev.ok),
                    score=float(obj.score),
                    alpha=alpha_val,
                    n_trades_closed=trades_val,
                    max_dd=dd_val,
                    penalty=penalty_val,
                    buy_hold_filo=buy_hold_val,
                    net_profit_strat=profit_val,
                    extras=getattr(ev.metrics, "extras", {}) or {},
                )

            def _forward_builder(
                    base_xlsx: Path,
                    selected_blocks: List[RegimeBlock],
                    out_xlsx: Path,
            ) -> None:

                out_xlsx = Path(out_xlsx)
                out_xlsx.parent.mkdir(parents=True, exist_ok=True)

                # Baseline vera S = {}
                if len(selected_blocks) == 0:
                    _write_forward_baseline_config(
                        base_config_path=base_xlsx,
                        out_config_path=out_xlsx,
                    )
                    return

                active_groups: List[str] = []
                seen = set()

                for block in selected_blocks:
                    regime_name = str(getattr(block, "regime", "") or "").strip()
                    if not regime_name:
                        continue
                    regime_up = regime_name.upper()
                    if regime_up in seen:
                        continue
                    seen.add(regime_up)
                    active_groups.append(regime_name)

                # Step 1: compose CONDITIONS from external template
                _write_forward_composed_conditions_config(
                    base_config_path=base_xlsx,
                    out_config_path=out_xlsx,
                    active_groups=active_groups,
                )

                # Step 2: merge trial-specific CONDITIONS overrides from selected per-regime best.xlsx files
                _merge_forward_conditions_from_blocks(
                    selected_blocks=selected_blocks,
                    out_config_path=out_xlsx,
                )

                try:
                    dbg_wb = load_workbook(out_xlsx, data_only=True)
                    if "CONDITIONS" in dbg_wb.sheetnames:
                        dbg_ws = dbg_wb["CONDITIONS"]
                        dbg_header = [c.value for c in dbg_ws[1]]
                        dbg_idx = {str(name): i + 1 for i, name in enumerate(dbg_header) if name is not None}
                        if "id" in dbg_idx and "rhs_value" in dbg_idx and "group" in dbg_idx:
                            for r in range(2, dbg_ws.max_row + 1):
                                row_id = dbg_ws.cell(r, dbg_idx["id"]).value
                                row_group = dbg_ws.cell(r, dbg_idx["group"]).value
                                if str(row_id) == "E_RA_RSI_LOW2" and str(row_group).strip().upper() == "G_RANGE":
                                    rhs_val = dbg_ws.cell(r, dbg_idx["rhs_value"]).value
                                    print(f"[DBG] composed candidate E_RA_RSI_LOW2 rhs_value={rhs_val}")
                                    break
                    dbg_wb.close()
                except Exception:
                    pass

                # Step 3: merge TUNING from selected per-regime best.xlsx files
                _merge_forward_tuning_from_blocks(
                    base_config_path=base_xlsx,
                    selected_blocks=selected_blocks,
                    out_config_path=out_xlsx,
                )



            def _forward_trade_reporter(eval_dir: Path, regime: str) -> None:
                return None

            selection_result = run_regime_forward_selection(
                input_csv=input_csv,
                timeframe=timeframe,
                base_config_xlsx=config_strategy,
                tuned_blocks=tuned_blocks,
                outdir=outdir,
                spec=fs_spec,
                evaluator=_forward_evaluator,
                builder=_forward_builder,
                trade_reporter=_forward_trade_reporter,
                seed=int(seed),
            )
    except Exception as e:
        (selection_dir / "selection_error.txt").write_text(str(e), encoding="utf-8")


    # ------------------------------------------------------------
    # Export final composed best blocks (CONDITIONS + TUNING)
    # ------------------------------------------------------------
    try:
        selected_groups_for_export: list[str] = []
        if isinstance(selection_result, dict):
            selected_groups = [
                str(x).strip()
                for x in (selection_result.get("selected_regimes") or [])
                if str(x).strip()
            ]
            joint_opt_groups = [
                str(x).strip()
                for x in (selection_result.get("joint_opt_regimes") or [])
                if str(x).strip()
            ]

            seen_groups: set[str] = set()
            selected_groups_for_export = []
            for g in selected_groups + joint_opt_groups:
                if g not in seen_groups:
                    seen_groups.add(g)
                    selected_groups_for_export.append(g)

        best_composed_xlsx = _resolve_best_composed_xlsx(outdir, selection_result)

        if best_composed_xlsx is not None and selected_groups_for_export:
            _export_multi_group_block_csvs_from_workbook(
                source_xlsx=best_composed_xlsx,
                base_tuning_xlsx=config_strategy,
                target_groups=selected_groups_for_export,
                out_dir=final_report_dir,
                conditions_filename="best_CONDITIONS.csv",
                tuning_filename="best_TUNING.csv",
            )
            print(
                "[DBG] composed export done "
                f"best_composed_xlsx={best_composed_xlsx} "
                f"groups={selected_groups_for_export}"
            )
        else:
            print(
                "[DBG] composed export skipped "
                f"best_composed_xlsx={best_composed_xlsx} "
                f"groups={selected_groups_for_export}"
            )
    except Exception as exc:
        print(f"[WARN] composed export best csv failed: {exc}")



    # Legacy per-regime summary (not the composed final report)
    summary_csv = outdir / "per_regime_summary.csv"

    standalone_profit_sum = 0.0
    standalone_trade_sum = 0
    standalone_has_profit = False

    for r in summary_rows:
        p = _to_float_maybe(r.get("profit"))
        n = _to_float_maybe(r.get("trade_count"))

        if p is not None:
            standalone_profit_sum += p
            standalone_has_profit = True

        if n is not None:
            standalone_trade_sum += int(n)

    overall_profit = standalone_profit_sum if standalone_has_profit else None
    overall_trades = standalone_trade_sum
    overall_ppt = _safe_profit_per_trade(overall_profit, overall_trades)

    # Keep composed-final metrics separate for the dedicated section below.
    if selection_result and isinstance(selection_result, dict):
        final_eval = selection_result.get("final_eval") or {}
        overall_buy_hold = final_eval.get("buy_hold_filo")
    else:
        final_eval = {}
        overall_buy_hold = None

    overall_alpha = (
        overall_profit - overall_buy_hold
        if overall_profit is not None and overall_buy_hold is not None
        else None
    )

    # Not meaningful as standalone aggregates in this summary row.
    overall_dd = None
    overall_score = None

    summary_rows_out = list(summary_rows)
    summary_rows_out.append(
        {
            "group": "OVERALL",
            "profit": overall_profit,
            "trade_count": overall_trades,
            "profit_per_trade": overall_ppt,
            "buy_hold_profit": overall_buy_hold,
            "alpha_vs_buyhold": overall_alpha,
            "max_drawdown": overall_dd,
            "score": overall_score,
            "standalone_hint": "",
            "best_xlsx": "",
        }
    )

    fieldnames = [
        "group",
        "profit",
        "trade_count",
        "profit_per_trade",
        "buy_hold_profit",
        "alpha_vs_buyhold",
        "max_drawdown",
        "score",
        "standalone_hint",
        "best_xlsx",
        "eval_dir",
        "gate_check_ok",
        "bad_entry_count",
    ]

    with summary_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, delimiter=";")
        w.writeheader()
        for row in summary_rows_out:
            w.writerow({k: _fmt_eu_num(v) for k, v in row.items()})

    # final console summary
    try:
        df_summary = pd.DataFrame(summary_rows_out)

        if not df_summary.empty:
            non_overall = df_summary[df_summary["group"] != "OVERALL"].copy()
            overall_only = df_summary[df_summary["group"] == "OVERALL"].copy()

            if "score" in non_overall.columns:
                non_overall = non_overall.sort_values(by="score", ascending=False, na_position="last")

            df_summary_print = pd.concat([non_overall, overall_only], ignore_index=True)

            printable_cols = [
                c for c in [
                    "group",
                    "profit",
                    "trade_count",
                    "profit_per_trade",
                    "buy_hold_profit",
                    "alpha_vs_buyhold",
                    "max_drawdown",
                    "score",
                    "standalone_hint",
                ]
                if c in df_summary_print.columns
            ]

            df_print = df_summary_print[printable_cols].copy()

            for col in [
                "profit",
                "profit_per_trade",
                "buy_hold_profit",
                "alpha_vs_buyhold",
                "max_drawdown",
                "score",
            ]:
                if col in df_print.columns:
                    df_print[col] = df_print[col].map(_fmt2)

            if "trade_count" in df_print.columns:
                df_print["trade_count"] = df_print["trade_count"].map(
                    lambda x: "" if pd.isna(x) else str(int(x))
                )

            print(
                "\n[NOTE] Le sezioni STANDALONE descrivono i blocchi regime isolati; "
                "le sezioni FINAL COMPOSED descrivono la strategia finale composta letta dal SIGNAL_* canonico."
            )
            print("\n=== REGIME SUMMARY (STANDALONE TUNED BLOCKS) ===")
            print(df_print.to_string(index=False))

            print("\n=== PROFIT / TRADE BY REGIME (STANDALONE) ===")
            ppt_cols = [c for c in ["group", "profit_per_trade", "trade_count", "profit"] if c in df_print.columns]
            print(df_print[ppt_cols].to_string(index=False))

        print("\n=== OVERALL STRATEGY VS BUY&HOLD (FINAL COMPOSED SIGNAL) ===")
        final_profit = final_eval.get("net_profit_strat") if isinstance(final_eval, dict) else None
        final_buy_hold = final_eval.get("buy_hold_filo") if isinstance(final_eval, dict) else None
        final_alpha = final_eval.get("alpha") if isinstance(final_eval, dict) else None
        final_trades = final_eval.get("n_trades_closed") if isinstance(final_eval, dict) else None
        final_dd = final_eval.get("max_dd") if isinstance(final_eval, dict) else None
        final_ppt = _safe_profit_per_trade(final_profit, final_trades)

        print(f"Strategy Profit : {_fmt2(final_profit)}")
        print(f"Buy&Hold Profit : {_fmt2(final_buy_hold)}")
        print(f"Alpha           : {_fmt2(final_alpha)}")
        print(f"Trade Count     : {'' if final_trades is None else int(final_trades)}")
        print(f"Profit/Trade    : {_fmt2(final_ppt)}")
        print(f"Max Drawdown    : {_fmt2(final_dd)}")
        print("\n=== SINGLE PROFIT / TRADE LIST (STANDALONE) ===")
        for _, r in df_summary_print.iterrows():
            group = r.get("group")
            if group == "OVERALL":
                continue

            ppt = _fmt2(r.get("profit_per_trade"))
            trades = r.get("trade_count")
            profit = _fmt2(r.get("profit"))
            trades_txt = "" if pd.isna(trades) else str(int(trades))

            print(f"- {group}: Profit/Trade={ppt} | Trades={trades_txt} | Profit={profit}")

        print("\n=== SINGLE TRADE PROFITS BY REGIME (STANDALONE) ===")
        for _, r in df_summary_print.iterrows():
            group = r.get("group")
            if group == "OVERALL":
                continue

            best_xlsx = r.get("best_xlsx")
            eval_dir = r.get("eval_dir")

            trade_values = _read_trade_profit_list_from_signal_dir(eval_dir)

            if not trade_values:
                trade_values = _read_trade_profit_list_from_best_xlsx(best_xlsx)

            if not trade_values:
                trade_values = _read_trade_profit_list_from_best_signal(best_xlsx)

            trade_list_txt = _fmt_trade_list(trade_values)

            expected_trades = r.get("trade_count")
            try:
                expected_trades_int = 0 if pd.isna(expected_trades) else int(expected_trades)
            except Exception:
                expected_trades_int = None

            actual_trades_int = len(trade_values)

            if trade_list_txt:
                if expected_trades_int is not None and expected_trades_int != actual_trades_int:
                    print(
                        f"- {group}: {trade_list_txt} "
                        f"[WARN count_mismatch summary={expected_trades_int} list={actual_trades_int}]"
                    )
                else:
                    print(f"- {group}: {trade_list_txt}")
            else:
                if expected_trades_int not in (None, 0):
                    print(
                        f"- {group}: (nessun trade) "
                        f"[WARN count_mismatch summary={expected_trades_int} list=0]"
                    )
                else:
                    print(f"- {group}: (nessun trade)")

        print("\n=== FINAL COMPOSED TRADE DETAIL ===")
        try:
            final_trade_rows = _read_final_composed_trade_rows(outdir, selection_result)
        except Exception:
            final_trade_rows = []

        if final_trade_rows:
            print(f"Trade count: {len(final_trade_rows)}")
            print("#trade; Profit del Trade; Regime apertura; Regime chiusura")

            for r in final_trade_rows:
                print(
                    f"{r['trade_no']}; "
                    f"{_fmt2(r['profit_per_trade'])}; "
                    f"{r['regime_open']}; "
                    f"{r['regime_close']}"
                )
        else:
            print("(no composed trades found)")

        print(f"\nSaved: {summary_csv}")
    except Exception as e:
        print(f"\n[WARN] Could not print regime summary table: {e}")

    return outdir
