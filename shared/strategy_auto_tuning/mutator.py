from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from shutil import copyfile
from typing import Any, Dict

from openpyxl import load_workbook

from .space import SearchSpace


def _norm_key(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _build_header_map(ws) -> Dict[str, int]:
    header: Dict[str, int] = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        key = _norm_key(cell.value)
        if key:
            header[key] = col_idx
    return header


def _build_row_map_by_id(ws, id_col_idx: int) -> Dict[str, int]:
    row_map: Dict[str, int] = {}
    for row_idx in range(2, ws.max_row + 1):
        key = _norm_key(ws.cell(row=row_idx, column=id_col_idx).value)
        if key:
            row_map[key] = row_idx
    return row_map


def _write_trial_by_cell_patch(
    base_xlsx: Path,
    out_xlsx: Path,
    mapping: Dict[str, tuple[str, str]],
    params: Dict[str, Any],
) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    copyfile(base_xlsx, out_xlsx)

    wb = load_workbook(out_xlsx)
    if "CONDITIONS" not in wb.sheetnames:
        raise ValueError("Config missing sheet: CONDITIONS")

    ws = wb["CONDITIONS"]
    header = _build_header_map(ws)

    if "id" not in header:
        raise ValueError("CONDITIONS missing required column: id")

    row_map = _build_row_map_by_id(ws, header["id"])

    for pname, pval in params.items():
        target = mapping.get(pname)
        if target is None:
            continue

        base_condition_id, field = target
        target_id = _norm_key(base_condition_id)
        target_field = _norm_key(field)

        if not target_id or not target_field:
            raise ValueError(
                f"Invalid mapping for param '{pname}': "
                f"base_condition_id={base_condition_id!r}, field={field!r}"
            )

        if target_field not in header:
            raise ValueError(
                f"CONDITIONS missing target field '{target_field}' "
                f"for param '{pname}' (base_condition_id={target_id})."
            )

        row_idx = row_map.get(target_id)
        if row_idx is None:
            raise ValueError(
                f"CONDITIONS has no row with id={target_id} "
                f"(required by param '{pname}')."
            )

        col_idx = header[target_field]
        ws.cell(row=row_idx, column=col_idx, value=pval)

    wb.save(out_xlsx)
    wb.close()


@dataclass
class ConfigMutator:
    """
    Apply sampled parameters to a config_strategy.xlsx by editing CONDITIONS.

    Mapping is defined by TUNING/SearchSpace columns:
      - base_condition_id (matches CONDITIONS.id)
      - field (target column in CONDITIONS)

    Important:
    this implementation patches only the target cells in CONDITIONS and keeps
    the original workbook untouched elsewhere. This avoids pandas full-sheet
    rewrites that can corrupt string cells such as operator '=='.
    """

    def apply(
        self,
        base_xlsx: Path,
        out_xlsx: Path,
        space: SearchSpace,
        params: Dict[str, Any],
    ) -> None:
        spec_by_name = space.as_dict()
        mapping: Dict[str, tuple[str, str]] = {}

        for pname, spec in spec_by_name.items():
            base_condition_id = getattr(spec, "base_condition_id", None)
            field = getattr(spec, "field", None)

            if base_condition_id is None or not field:
                continue

            mapping[pname] = (_norm_key(base_condition_id), _norm_key(field))

        _write_trial_by_cell_patch(
            base_xlsx=Path(base_xlsx),
            out_xlsx=Path(out_xlsx),
            mapping=mapping,
            params=params,
        )


def write_trial_config(
    base_config_path: str | Path,
    out_config_path: str | Path,
    params: Dict[str, Any],
) -> None:
    """
    Engine-compatible writer.

    Reads mapping from sheet TUNING:
      - param_name
      - base_condition_id   (string, matches CONDITIONS.id)
      - field               (target column in CONDITIONS)

    Then copies the original XLSX and patches only the target cells in sheet
    CONDITIONS, preserving every other cell exactly as in the base workbook.
    """
    from openpyxl import load_workbook

    base_config_path = Path(base_config_path)
    out_config_path = Path(out_config_path)
    out_config_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(base_config_path, data_only=False)
    if "TUNING" not in wb.sheetnames:
        raise ValueError("Config missing sheet: TUNING")
    if "CONDITIONS" not in wb.sheetnames:
        raise ValueError("Config missing sheet: CONDITIONS")

    ws_tuning = wb["TUNING"]
    tuning_header = _build_header_map(ws_tuning)

    for c in ["param_name", "base_condition_id", "field"]:
        if c not in tuning_header:
            wb.close()
            raise ValueError(f"TUNING missing required column: {c}")

    mapping: Dict[str, tuple[str, str]] = {}
    for row_idx in range(2, ws_tuning.max_row + 1):
        pname = _norm_key(ws_tuning.cell(row=row_idx, column=tuning_header["param_name"]).value)
        base_id = _norm_key(ws_tuning.cell(row=row_idx, column=tuning_header["base_condition_id"]).value)
        field = _norm_key(ws_tuning.cell(row=row_idx, column=tuning_header["field"]).value)

        if not pname or pname.upper() == "USAGE":
            continue
        if not base_id or not field:
            continue

        mapping[pname] = (base_id, field)

    wb.close()

    _write_trial_by_cell_patch(
        base_xlsx=base_config_path,
        out_xlsx=out_config_path,
        mapping=mapping,
        params=params,
    )