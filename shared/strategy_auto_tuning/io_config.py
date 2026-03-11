# shared/strategy_auto_tuning/io_config.py
"""
Strategy Auto-Tuning — Excel I/O (config_strategy.xlsx)

Responsibilities (v1):
- Read the "TUNING" sheet into a pandas DataFrame.
- Write a trial config .xlsx by copying the base workbook and replacing ONLY the "TUNING" sheet.

Design goals:
- Preserve all other sheets: CONDITIONS, ENUMS, KPI_COLUMNS, ...
- Robust numeric parsing delegated to space.py; here we keep raw cell values where possible.
- Avoid brittle cell-by-cell edits on other sheets.

Notes:
- We intentionally rewrite the whole "TUNING" sheet to keep behavior deterministic.
- Formatting of "TUNING" may not be preserved. This is acceptable for a machine-driven config.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import shutil

DEFAULT_TUNING_SHEET = "TUNING"


def read_tuning_sheet(
    xlsx_path: str | Path,
    *,
    sheet_name: str = DEFAULT_TUNING_SHEET,
) -> pd.DataFrame:
    """
    Read the TUNING sheet as DataFrame.

    Returns a DataFrame with the sheet columns as-is (no normalization here).
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(str(xlsx_path))

    try:
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, engine="openpyxl")
    except ValueError as e:
        # sheet not found or similar
        raise ValueError(f"Missing sheet '{sheet_name}' in {xlsx_path.name}") from e

    if df is None:
        raise ValueError(f"Failed to read sheet '{sheet_name}' from {xlsx_path.name}")

    # Keep empty strings instead of NaN for stability in downstream logs
    df = df.copy()
    return df


def write_tuning_sheet(
    base_xlsx_path: str | Path,
    tuning_df: pd.DataFrame,
    out_xlsx_path: str | Path,
    *,
    sheet_name: str = DEFAULT_TUNING_SHEET,
    logger: Optional[Any] = None,
) -> None:
    """
    Create a new XLSX by copying base workbook and replacing ONLY the tuning sheet.

    Parameters
    ----------
    base_xlsx_path:
        Path to the base config_strategy.xlsx (template).
    tuning_df:
        DataFrame for the tuning sheet content (must include headers).
    out_xlsx_path:
        Output xlsx path (will be overwritten).
    sheet_name:
        Name of the tuning sheet (default "TUNING").
    """
    base_xlsx_path = Path(base_xlsx_path)
    out_xlsx_path = Path(out_xlsx_path)

    if not base_xlsx_path.exists():
        raise FileNotFoundError(str(base_xlsx_path))

    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(base_xlsx_path)

    # Remove existing tuning sheet if present
    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)

    # Create a new tuning sheet at the end (or you can insert at original index if you prefer)
    ws_new = wb.create_sheet(sheet_name)

    _write_dataframe_to_sheet(ws_new, tuning_df)

    # Optional: freeze header row for readability (harmless)
    ws_new.freeze_panes = "A2"

    # Save (overwrite)
    wb.save(out_xlsx_path)

    if logger is not None and hasattr(logger, "info"):
        logger.info(f"[IO] wrote trial config: {out_xlsx_path}")
    else:
        # keep silent by default (engine can log)
        pass

def build_regime_wise_config_v3(
    base_xlsx_path: str | Path,
    out_xlsx_path: str | Path,
    *,
    target_group: str,
    conditions_sheet_name: str = "CONDITIONS",
    logger: Optional[Any] = None,
) -> Path:
    """
    Build a temporary config_strategy workbook for regime-wise v3 evaluation.

    Semantics:
    - full dataset is kept unchanged outside this function
    - only CONDITIONS sheet is edited
    - rows with group == target_group are enabled
    - rows with any other non-empty group are disabled
    - blank group rows are left unchanged
    """
    base_xlsx_path = Path(base_xlsx_path)
    out_xlsx_path = Path(out_xlsx_path)

    if not base_xlsx_path.exists():
        raise FileNotFoundError(str(base_xlsx_path))

    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base_xlsx_path, out_xlsx_path)

    wb = load_workbook(out_xlsx_path)

    if conditions_sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Missing sheet '{conditions_sheet_name}' in {out_xlsx_path.name}"
        )

    ws = wb[conditions_sheet_name]

    header = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]
    header_norm = {str(v).strip().lower(): j for j, v in enumerate(header, start=1) if v is not None}

    if "group" not in header_norm:
        raise ValueError(
            f"Missing required column 'group' in sheet '{conditions_sheet_name}'"
        )
    if "enabled" not in header_norm:
        raise ValueError(
            f"Missing required column 'enabled' in sheet '{conditions_sheet_name}'"
        )

    col_group = header_norm["group"]
    col_enabled = header_norm["enabled"]

    target_group_norm = str(target_group).strip()

    for i in range(2, ws.max_row + 1):
        group_val = ws.cell(row=i, column=col_group).value
        group_str = "" if group_val is None else str(group_val).strip()

        # Leave rows without group untouched
        if not group_str:
            continue

        ws.cell(row=i, column=col_enabled, value=(group_str == target_group_norm))

    wb.save(out_xlsx_path)

    if logger is not None and hasattr(logger, "info"):
        logger.info(
            f"[IO] wrote regime-wise v3 config: {out_xlsx_path} "
            f"(target_group={target_group_norm})"
        )

    return out_xlsx_path


def _write_dataframe_to_sheet(ws: Worksheet, df: pd.DataFrame) -> None:
    """
    Write DataFrame into openpyxl worksheet.

    Behavior:
    - Writes headers in row 1
    - Writes values row-by-row
    - Converts NaN to empty string
    """
    if df is None:
        raise ValueError("tuning_df is None")

    # Ensure stable column order
    cols = list(df.columns)

    # Header
    for j, c in enumerate(cols, start=1):
        ws.cell(row=1, column=j, value=str(c))

    # Body
    for i in range(len(df)):
        for j, c in enumerate(cols, start=1):
            v = df.iloc[i][c]
            if pd.isna(v):
                v = ""
            ws.cell(row=i + 2, column=j, value=v)